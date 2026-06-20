from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import models


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _append_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 40)


def _save_workbook(wb: Workbook) -> Path:
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    path = Path(tmp.name)
    wb.save(path)
    return path


def _get_visitor_year(db: Session, academic_year: int | None = None, year_id=None) -> models.VisitorSchoolYear:
    query = db.query(models.VisitorSchoolYear)
    if year_id is not None:
        year = query.filter(models.VisitorSchoolYear.id == year_id).first()
    elif academic_year is not None:
        year = query.filter(models.VisitorSchoolYear.academic_year == academic_year).first()
    else:
        year = None
    if not year:
        raise ValueError("visitor_year_not_found")
    return year


def _safe_sheet_title(value: str) -> str:
    # Excel sheet titles are limited to 31 characters.
    return value[:31]


def _style_range(ws, cell_range: str, *, fill=None, font=None, border=None, alignment=None, number_format: str | None = None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment
            if number_format is not None:
                cell.number_format = number_format


def _period_formula_ranges(year: models.VisitorSchoolYear, periods: list[models.VisitorPeriod], period_types: set[models.VisitorPeriodType]) -> list[str]:
    months = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2]
    month_columns = {month: get_column_letter(index + 2) for index, month in enumerate(months)}
    ranges: list[str] = []
    for period in periods:
        if period.period_type not in period_types or not period.start_date or not period.end_date:
            continue
        for month in months:
            calendar_year = year.academic_year if month >= 3 else year.academic_year + 1
            month_start = date(calendar_year, month, 1)
            if month == 12:
                month_end = date(calendar_year, 12, 31)
            else:
                next_month_year = calendar_year + 1 if month == 12 else calendar_year
                next_month = 1 if month == 12 else month + 1
                month_end = date(next_month_year, next_month, 1) - timedelta(days=1)
            start = max(period.start_date, month_start, year.start_date)
            end = min(period.end_date, month_end, year.end_date)
            if start > end:
                continue
            column = month_columns[month]
            start_row = start.day + 3
            end_row = end.day + 3
            ranges.append(f"{column}{start_row}:{column}{end_row}" if start_row != end_row else f"{column}{start_row}")
    return ranges


def _sum_formula(ranges: list[str]) -> str:
    return f"=SUM({','.join(ranges)})" if ranges else "=0"


def build_visitors_excel(db: Session, academic_year: int | None = None, *, year_id=None) -> Path:
    year = _get_visitor_year(db, academic_year=academic_year, year_id=year_id)

    wb = Workbook()
    ws = wb.active
    title = f"{year.academic_year}학년도 참고열람실 출입자 통계"
    ws.title = _safe_sheet_title(title)

    # Legacy worksheet shape: A1:V43, with A:M as the monthly calendar,
    # F:K as bottom summary, and O:R as the counter continuation helper.
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    helper_fill = PatternFill("solid", fgColor="E2F0D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)
    number_format = "#,##0"

    ws.merge_cells("A1:K1")
    ws["A1"] = title
    _style_range(ws, "A1:K1", fill=white_fill, font=title_font, border=border, alignment=center)

    months = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2]
    month_columns = {month: index + 2 for index, month in enumerate(months)}

    for index, month in enumerate(months, start=2):
        cell = ws.cell(row=3, column=index, value=f"{month}월")
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for day in range(1, 32):
        cell = ws.cell(row=day + 3, column=1, value=f"{day}일")
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        for column in range(2, 14):
            data_cell = ws.cell(row=day + 3, column=column)
            data_cell.border = border
            data_cell.alignment = center
            data_cell.number_format = number_format

    entries = (
        db.query(models.VisitorDailyCount)
        .filter(models.VisitorDailyCount.school_year_id == year.id)
        .all()
    )
    for entry in entries:
        visit_date = entry.visit_date
        if visit_date.month not in month_columns:
            continue
        expected_year = year.academic_year if visit_date.month >= 3 else year.academic_year + 1
        if visit_date.year != expected_year or not (1 <= visit_date.day <= 31):
            continue
        ws.cell(row=visit_date.day + 3, column=month_columns[visit_date.month], value=entry.daily_visitors)

    ws["A35"] = "합계"
    ws["A36"] = "개관일수"
    for column in range(2, 14):
        column_letter = get_column_letter(column)
        ws.cell(row=35, column=column, value=f"=SUM({column_letter}4:{column_letter}34)")
        ws.cell(row=36, column=column, value=f"=COUNTA({column_letter}4:{column_letter}34)")
    ws["N35"] = "=SUM(B35:M35)"
    ws["N36"] = "=SUM(B36:M36)"
    _style_range(ws, "A35:N36", fill=summary_fill, font=bold_font, border=border, alignment=center, number_format=number_format)

    periods = (
        db.query(models.VisitorPeriod)
        .filter(models.VisitorPeriod.school_year_id == year.id)
        .all()
    )
    semester_ranges = _period_formula_ranges(
        year,
        periods,
        {models.VisitorPeriodType.SEMESTER_1, models.VisitorPeriodType.SEMESTER_2},
    )
    break_ranges = _period_formula_ranges(
        year,
        periods,
        {models.VisitorPeriodType.SUMMER_BREAK, models.VisitorPeriodType.WINTER_BREAK},
    )

    ws.merge_cells("G38:H38")
    ws.merge_cells("J38:K38")
    ws.merge_cells("G39:H39")
    ws.merge_cells("J39:K39")
    ws["F38"] = "개관일수"
    ws["G38"] = "=SUM(B36:M36)"
    ws["I38"] = "총출입자수"
    ws["J38"] = "=SUM(B35:M35)"
    ws["F39"] = "학기중"
    ws["G39"] = _sum_formula(semester_ranges)
    ws["I39"] = "방학중"
    ws["J39"] = _sum_formula(break_ranges)
    _style_range(ws, "F38:K39", fill=summary_fill, font=bold_font, border=border, alignment=center, number_format=number_format)

    running = (
        db.query(models.VisitorRunningTotal)
        .filter(models.VisitorRunningTotal.school_year_id == year.id)
        .first()
    )
    next_previous_total = running.previous_total if running and running.previous_total is not None else None
    if next_previous_total is None:
        latest_with_current = (
            db.query(models.VisitorDailyCount)
            .filter(
                models.VisitorDailyCount.school_year_id == year.id,
                models.VisitorDailyCount.current_total.isnot(None),
            )
            .order_by(models.VisitorDailyCount.visit_date.desc())
            .first()
        )
        if latest_with_current:
            next_previous_total = latest_with_current.current_total

    helper_headers = {
        "O3": "전일 합",
        "O6": "Count1",
        "P6": "Count2",
        "Q6": "금일 합",
        "R6": "금일 출입자",
    }
    for cell_ref, value in helper_headers.items():
        ws[cell_ref] = value
    ws["O4"] = next_previous_total
    ws["Q7"] = '=IF(OR(O7="",P7=""),"",O7+P7)'
    ws["R7"] = '=IF(OR(Q7="",O4=""),"",Q7-O4)'
    _style_range(ws, "O3:R7", fill=white_fill, border=border, alignment=center, number_format=number_format)
    _style_range(ws, "O3:R3", fill=helper_fill, font=bold_font, border=border, alignment=center)
    _style_range(ws, "O6:R6", fill=helper_fill, font=bold_font, border=border, alignment=center)

    widths = {
        "A": 8,
        "B": 9,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 9,
        "M": 9,
        "N": 11,
        "O": 13,
        "P": 13,
        "Q": 13,
        "R": 15,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8
    for row in range(3, 44):
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "B4"
    return _save_workbook(wb)


def build_serials_excel(db: Session) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    publications = db.query(models.SerialPublication).order_by(models.SerialPublication.title.asc()).all()
    _append_sheet(
        wb,
        "간행물 목록",
        ["제목", "ISSN", "수집 유형", "서가 구역", "서가 ID", "행", "열", "행 끝", "열 끝", "서가 메모", "비고"],
        [
            [
                item.title,
                item.issn or "",
                item.acquisition_type.value,
                item.shelf_section,
                str(item.shelf_id) if item.shelf_id else "",
                item.shelf_row,
                item.shelf_column,
                item.shelf_row_end,
                item.shelf_column_end,
                item.shelf_note or "",
                item.remark or "",
            ]
            for item in publications
        ],
    )

    shelves = db.query(models.SerialShelf).order_by(models.SerialShelf.code.asc()).all()
    _append_sheet(
        wb,
        "서가 목록",
        ["코드", "배치도 ID", "서가 타입 ID", "X", "Y", "회전", "메모"],
        [[item.code, str(item.layout_id), str(item.shelf_type_id), item.x, item.y, item.rotation, item.note or ""] for item in shelves],
    )

    layouts = db.query(models.SerialLayout).order_by(models.SerialLayout.name.asc()).all()
    _append_sheet(
        wb,
        "배치도 목록",
        ["이름", "너비", "높이", "메모", "벽 데이터"],
        [[item.name, item.width, item.height, item.note or "", str(item.walls or "")] for item in layouts],
    )

    shelf_types = db.query(models.SerialShelfType).order_by(models.SerialShelfType.name.asc()).all()
    _append_sheet(
        wb,
        "서가 타입",
        ["이름", "너비", "높이", "행", "열", "색상", "메모"],
        [[item.name, item.width, item.height, item.rows, item.columns, item.color or "", item.note or ""] for item in shelf_types],
    )
    return _save_workbook(wb)
